import os
from test import start_test_block, handle_choice, send_question, \
    test_is_active, handle_text_answer, question_is_open
from contacts import send_contact_request, set_contacts, user_is_leaving_contacts
from decorators import data_handler
from telebot import TeleBot, types
#import openpyxl
import re
import traceback
import psycopg2
import json
import xlsxwriter
import openpyxl
from openpyxl import Workbook
import io

import asyncio
import threading
from time import sleep
import sys
from datetime import date, datetime
from datetime import timedelta
from json_to_posgresql import Json_to_posgresql
from contact import Contact
from bot_config import Bot_config


dir_path = os.path.dirname(os.path.realpath(__file__))
bot_config=Bot_config()
#comand admins
admins_ids = bot_config.admins_ids
token = bot_config.token
bot = TeleBot(token)


# Обработчик команды /start
@bot.message_handler(commands=['start'])
@data_handler
def start(cursor, message):     
    if message.from_user.id not in admins_ids:
        check_if_user_exists = f'SELECT user_id FROM users_contacts WHERE user_id = {message.from_user.id}'
        cursor.execute(check_if_user_exists)
        user_exists = cursor.fetchone()

        if user_exists is not None:
            return

    set_contacts(message.from_user.id, message.from_user.username, True)

    keyboard = types.InlineKeyboardMarkup()

    start_button = types.InlineKeyboardButton('Согласен', callback_data='leave_contacts')
    keyboard.add(start_button)

    bot_image_path = f'{dir_path}//..//resources//томка1.png'
    with open(bot_image_path, 'rb') as image:
        bot.send_photo(message.chat.id, image.read())

    start_message = 'Я чат-бот Томка, помогу определить, насколько прокачено у тебя логическое мышление,' \
                    ' ведь на большинстве ИТ вакансий оно пригодится. Если ответишь правильно на все вопросы,' \
                    ' получишь приз. Также я познакомлю тебя с актуальными позициями в ТМК++ и буду держать кулачки,' \
                    ' чтобы ты стал частью нашей команды.\nДля продолжения общения, ' \
                    'необходимо твое согласие на обработку персональных данных, если не возражаешь,' \
                    ' то ознакомься с текстом согласия и Политикой Компании в области защиты персональных данных' \
                    ' и после этого нажми на кнопку «Согласен». ' \
                    'Текст согласия ты сможешь найти во вложениях ниже:'

    bot.send_message(message.chat.id, start_message, reply_markup=keyboard)

    agreement_path = f'{dir_path}//..//resources//согласие.docx'
    with open(agreement_path, 'rb') as doc:
        bot.send_document(
            message.chat.id,
            doc.read(),
            visible_file_name='Согласие на обработку персональных данных.docx'
        )



# Обработчик команды /data
@bot.message_handler(commands=['update_DB'])
@data_handler
def update_DB(cursor, message):
    if ((message.chat.type=="private") and (message.from_user.id  in admins_ids)):
        text = Json_to_posgresql.update_DB(cursor)
        bot.send_message(message.from_user.id,text)
        print(text)
    
    
   
# Обработчик команды /data
@bot.message_handler(commands = ['get_test_results'])
@data_handler
def get_test_results(cursor, message):
    print('get_test_results')
    if ((message.chat.type == "private") and (message.from_user.id in admins_ids)):        
        try:
            #create exel
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = f'Участники'

            # Заполняем ячейки в примере
            columns_index = 1
            #for columns leters
            columns_leters = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z','aa','ab','ac','ad','ae','af','ag','ah','ai','aj','ak','al','am','an','ao','ap','aq','ar','as','at','au','av','aw','ax','ay','az']
            #read estructure json
            with open(f'{dir_path}//..//data//contacts_structure.json', encoding = 'utf-8') as cs:
                        contacts_structure = json.loads(cs.read())
        
            row_index = 1
            columns_index = 0
            list_contact = []
            list_contact.append(Contact(columns_index,"Telegram"))
            position = f'{columns_leters[columns_index]}{row_index}'
            sheet[position] = "Telegram"
            columns_index += 1
            position = f'{columns_leters[columns_index]}{row_index}'
            sheet[position] = "Правильных ответов"
            columns_index += 1
    
            for contact_str in contacts_structure['elements']:
                c = Contact(columns_index,contact_str['name'])
                list_contact.append(c)
                position = f'{columns_leters[columns_index]}{row_index}'
                sheet[position] = contact_str['name']
                columns_index += 1
            #nex row
            row_index += 1
            # now the data
            get_all_user=f'''SELECT user_id, is_test_active, block_id, questions_order, question_index, contact_index, agreement, entry_date FROM users;'''
            cursor.execute(get_all_user)
            all_user = cursor.fetchall()
            for row_user in all_user:
                get_all_contacts = f'SELECT user_id, contact_name, contact_value FROM users_contacts WHERE user_id={row_user[0]} '
                cursor.execute(get_all_contacts)
                all_contacts = cursor.fetchall()
                for row_contacts in all_contacts:
                    for c in list_contact:
                        if(c.name == row_contacts[1]):
                            #write de data en exel
                            position = f'{columns_leters[c.id]}{row_index}'
                            sheet[position] = f'{row_contacts[2]}'
                            break
                get_count_users_answers = f'''SELECT COUNT (is_correct) FROM users_answers    WHERE user_id={row_user[0]} and is_correct = {1} ;  '''
                cursor.execute(get_count_users_answers)
                count_users_answers = cursor.fetchone()
                position = f'{columns_leters[1]}{row_index}'
                sheet[position] = f'{count_users_answers[0]}' 
                #nex row
                row_index += 1

            # Создаем буфер и записываем в него файл Excel
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            # Отправляем файл в чат
            bot.send_document(message.chat.id, ('Результаты тестирования.xlsx', excel_buffer))
        except (Exception) as error:
            print("Error while connecting to PostgreSQL: ", error)
            return False
        return True
    

    #get_all_contacts = f'SELECT user_id, contact_name, contact_value FROM users_contacts'
    #all_contacts = cursor.execute(get_all_contacts).fetchall()    
    #all_contacts = pd.DataFrame(all_contacts, columns=['user_id', 'contact_name', 'contact_value'])
    #all_contacts_df = pd.DataFrame()

    # Нужно доработать фукнцию!

    #get_all_contacts = f'SELECT user_id, contact_name, contact_value FROM users_contacts'
    #all_contacts = cursor.execute(get_all_contacts).fetchall()
    #all_contacts = pd.DataFrame(all_contacts, columns=['user_id', 'contact_name', 'contact_value'])
    #all_contacts_df = pd.DataFrame()


    #get_users_answers = f'SELECT * from users_answers'
    #users_answers = cursor.execute(get_users_answers).fetchall()
    #users_answers = pd.DataFrame(users_answers, columns=['user_id', 'question_id', 'is_correct'])

    #for user_id, group in all_contacts.groupby('user_id'):
    #    index = len(all_contacts_df)

    #    for row_index, row in group.iterrows():
    #        all_contacts_df.loc[index, row.contact_name] = row.contact_value

    #    correct_answers_count = users_answers[
    #        (users_answers['user_id'] == user_id)
    #        & users_answers['is_correct'] == 1
    #    ].shape[0]

    #    all_contacts_df.loc[index, 'Правильных ответов'] = correct_answers_count

    #table_path = f'{dir_path}//..//tables//Результаты тестирования.xlsx'

    #with pd.ExcelWriter(table_path) as writer:
    #    if len(all_contacts_df) > 0:
    #        all_contacts_df.to_excel(writer, sheet_name='Участники', index=False)

    #wb = openpyxl.load_workbook(filename=table_path)
    #for sheet_name in wb.sheetnames:
    #    worksheet = wb[sheet_name]
    #    for col in worksheet.columns:
    #        max_length = 0
    #        column = col[0].column_letter
    #        for cell in col:
    #            try:
    #                if len(str(cell.value)) > max_length:
    #                    max_length = len(str(cell.value))
    #            except:
    #                pass
    #        adjusted_width = (max_length + 2) * 1.2
    #        worksheet.column_dimensions[column].width = adjusted_width

    #wb.save(table_path)

    #with open(table_path, 'rb') as table:
    #    bot.send_document(message.chat.id, table.read(), visible_file_name=table_path[table_path.rfind('/'):])
    #    table.close()


# Обработчик любого введённого текста
@bot.message_handler(content_types=['text'])
def handle_text(message):
    user_id = str(message.from_user.id)
    chat_id = message.chat.id
    message_text = message.text    
    if user_is_leaving_contacts(user_id):
        message_text = re.sub(r'[^\w\s\.\+\-]', '', message_text)
        if set_contacts(user_id, message_text):
            send_contact_request(chat_id, user_id)
        else:
            bot.send_message(chat_id, 'Неправильный формат, попробуй ещё раз')
    elif test_is_active(user_id) and question_is_open(user_id):
        handle_text_answer(chat_id, user_id, message_text)
        send_question(chat_id, user_id)
    else:
        bot.send_message(chat_id, message_text)


# Обработчик всех кнопок с 'test' в содержимом callbackQuery.data
@bot.callback_query_handler(lambda query: 'test' in query.data)
def handle_test(callback):
    user_id = str(callback.from_user.id)
    chat_id = callback.message.chat.id
    if callback.data == 'test_start':
        bot.delete_message(chat_id, callback.message.id)        
        if test_is_active(user_id):
            bot.send_message(chat_id, 'Тест уже идёт')
        else:
            start_test_block(user_id)
            send_question(chat_id, user_id)
    elif callback.data == 'test_continue':
        bot.delete_message(chat_id, callback.message.id)
        start_test_block(user_id)
        send_question(chat_id, user_id)
    else:
        handle_choice(chat_id, user_id, callback)
        send_question(chat_id, user_id)


# Обработчик всех кнопок с 'contacts' в содержимом callbackQuery.data
@bot.callback_query_handler(lambda query: 'contacts' in query.data)
@data_handler
def handle_contacts(cursor, callback):
    bot.answer_callback_query(callback.id)
    user_id = str(callback.from_user.id)
    chat_id = callback.message.chat.id
    if callback.data == 'leave_contacts':
        send_contact_request(chat_id, user_id, True)


@data_handler
def database_cleaner(cursor):
    print ("database_cleaner")    
    try:
        #cursor = connection.cursor()
        today = datetime.today().date()      
        day_delta=timedelta(days=60)
        old_day=today-day_delta
        value_date = old_day.strftime('%Y-%m-%d')
        print(value_date)               
        get_user_id = f'''SELECT user_id FROM users WHERE entry_date <= '{value_date}';'''
        cursor.execute(get_user_id)
        users_id = cursor.fetchall()
        print("----")
        print(users_id)
        print("----")
        for row in users_id:
            delete_users_answers = "DELETE FROM users_answers WHERE user_id = "+str(row[0])+";"
            delete_users_contacts ="DELETE FROM users_contacts WHERE user_id = "+str(+row[0])+";"
            delete_users_open_answers ="DELETE FROM users_open_answers	WHERE user_id = "+str(+row[0])+";"
            delete_user = "DELETE FROM users WHERE user_id = '"+str(+row[0])+"';"
            delete_data = delete_users_answers+delete_users_contacts+delete_users_open_answers+delete_user
            cursor.execute(delete_data)          
    except (Exception, psycopg2.Error) as error:
        print("Error while connecting to PostgreSQL: ", error)
    

def database_cleaner_loop():
    print("database_cleaner_loop is running")
    while True:
        try:
            sleep(10)                    
            database_cleaner()
            sleep(86400)         
        except Exception as e:
            print("*************")
            print(datetime.datetime.now(), e)
            print("")
            continue
        

    
#threading.Thread(target=starBot).start()
threading.Thread(target=database_cleaner_loop).start()
# Запуск бота
if __name__ == '__main__':    
    while True:
        try:
           print("Запускаем бота")
           bot.polling(none_stop=True, interval=0)           
        except Exception as e:
            print("*************")
            print(datetime.datetime.now(), e)
            print("")
            print(traceback.format_exc())
            print("*************")
            sleep(5)
            continue

