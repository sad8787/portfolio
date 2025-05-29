import telebot
from telebot import types
from flask import Flask, request, redirect
import requests
import json
import xlsxwriter
import openpyxl
from openpyxl import Workbook
import io

import threading
import os
import configparser
import datetime 
from datetime import datetime, timedelta
from time import sleep
from user import User
from employee import Employee

from user_status import User_status
from user_chat_group import User_chat_group
from decorators import data_handler
from db_utils import db_utils_static
from bot_utils import bot_utils_static
from sso_params import SSO_params


##config
dir_path = os.path.dirname(os.path.realpath(__file__))
config = configparser.ConfigParser()
config.read(f'{dir_path}//..//data//config.ini')

# Конфигурация OpenID Connect
auth_endpoint =config['url']['auth_endpoint'] #'YOUR_OPENID_CONNECT_AUTH_ENDPOINT'
token_endpoint = config['url']['token_endpoint'] #'YOUR_OPENID_CONNECT_TOKEN_ENDPOINT'
userinfo_endpoint = config['url']['userinfo_endpoint'] #'YOUR_OPENID_CONNECT_USERINFO_ENDPOINT'
redirect_uri = config['url']['redirect_uri']#'YOUR_REDIRECT_URI'
token_introspect=config['url']['token_introspect']# token_introspect  чтобы проверить, активен ли токен

url_get_employee = config['url']['url_get_employee']# url for get employee
token_get_employee = config['url']['token_get_employee']# token for get employee
#token_get_employee = os.environ.get('token_get_employee')# token for get employee


client_id = config['url']['client_id'] #'YOUR_OPENID_CONNECT_CLIENT_ID'
client_secret = config['url']['client_secret'] #'YOUR_OPENID_CONNECT_CLIENT_SECRET'
#client_id = os.environ.get('client_id') #'YOUR_OPENID_CONNECT_CLIENT_ID'
#client_secret = os.environ.get('client_secret') #'YOUR_OPENID_CONNECT_CLIENT_SECRET'

sso_params = SSO_params(auth_endpoint, token_endpoint, userinfo_endpoint, redirect_uri, token_introspect, client_id, client_secret)

host_flask=config['web']['host']
port_flask=config['web']['port']

# Token
bot_token = config['bot']['token']
#bot_token = os.environ.get('bot_token') #bot token

#url_bot
bot_url = config['bot']['bot_url']


#comand admins
comands_admins = [1555620161, 1371591132]

#id bot
bot_id = [6281036303, 6403728414]
#main_group_chat = -1001312443368

count_days_for_ban = 3
count_days_for_ban_for_ever = 30
count_days_for_ban_in_tlg = 50
cleaning_period = 86400
message_text = f'''Мы уведомляем вас, что вы должны подтвердить свои полномочия через нашего бота {bot_url}'''
help_message_for_group = f'''Данная группа используется работниками ТМК и с помощью нашего бота {bot_url} мы идентифицируем и проверяем участников группы.'''
help_message_in_bot_for_user = f'''Этот бот создан для прохождения аутентификации пользователей ТМК в TMK ID.
/start это команда для начала работы с ботом.
Для прохождения аутентификации нажимаем кнопку "Аутентификация" и подтверждаем свои полномочия в ТМК ID.'''
help_message_in_bot_for_admin = f'''Этот бот создан для прохождения аутентификации пользователей ТМК в TMK ID.
    /start это команда для начала работы с ботом.
    Для прохождения аутентификации нажимаем кнопку "Аутентификация" и подтверждаем свои полномочия в ТМК ID.
    /get_admins команда, которая запускается только в группе если бота добавили уже в существующую группу. Мы получаем список всех администраторов и при необходимости оправляем запрос на аутентификацию.
    /help команда для получения информации    
    /invitation командя для бота (не работает в группе), которая создает пригласительные ссылки для групп где присутствует бот.
    /list_user_excel командя для бота (не работает в группе), которая создает Excel файл. На первом листе файла все пользователи бота с необходимой информацией. Следующие листы создаются с теми группами где есть бот и списком пользователей группы.
    /promote_on_service-telegram_id командя для бота (не работает в группе), которая устанавливает для польователя признак is_on_service=True (дежурный администратор) и его мы не просим проходить аутентификацию.
    /downgrade_on_service-telegram_id командя для бота (не работает в группе), которая для польователя нимает признак дежурного is_on_service=False (НЕ дежурный администратор) и его мы будем просим проходить аутентификацию.'''

bot = telebot.TeleBot(bot_token)

# Создаем Flask-приложение для обработки аутентификации
app = Flask(__name__)
 
#для telebot BOT
# Запускаем бота
def starBot():
    while True:
        print(f'{datetime.today()} Запускаем бота')
        try:           
            bot.polling(none_stop=True, interval=0)   
        except (Exception)as error:
            print(f'{datetime.today()} ERROR in bot.polling():  description: {error}  ')
            continue
        
    

# Обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    print("start comand")
    text=message.text
    print(message.chat.id)
    if(bot_utils_static.is_group_chat(message)):
        return  False
    
    if(text != "/start"):
        code = "-" + message.text.split("-")[1]   
        print(f'start code {code}')
        bot_utils_static.create_button_auth(message.from_user.id, sso_params, bot, id_chat_group=code)
    else:       
        bot_utils_static.create_button_auth(message.from_user.id, sso_params, bot)


# Команда, которая запускается только в группе если бота добавили уже в существующую группу. Мы получаем список всех администраторов и при необходимости оправляем запрос на аутентификацию.
@bot.message_handler(commands=['get_admins'])
@data_handler
def getAdmin(cursor, message):     
    print("get_admins")
    print(f'{message.from_user.id} sey {message.text} in {message.chat.id}' )
    try:
        if(bot_utils_static.is_not_admin(message, comands_admins)):
            return False  
        if(bot_utils_static.is_private_chat(message)):     
            return False
        bot.delete_message(message.chat.id, message.message_id)

        admins = bot.get_chat_administrators(message.chat.id)
        for admin in admins:
            print(admin.user)
            telegram_user_id = admin.user.id

            user = User(telegram_user_id, User_status.not_confirmed, admin.user.is_bot)
            current_user = db_utils_static.check_user_exists(cursor, user)           

            user_chat_group = User_chat_group(telegram_user_id, message.chat.id)
            db_utils_static.check_user_chat_group_exists(cursor, user_chat_group)
            
            if current_user.status != User_status.confirmed.value:
                bot_utils_static.create_button_auth(telegram_user_id, sso_params, bot)  
            else:
                bot_utils_static.set_chat_administrator_custom_title(cursor, user, bot)
        bot.send_message(message.chat.id, message_text)
    except (Exception)as error:
        print(f'{datetime.today()} ERROR in getAdmin()   description: {error}  ')




   


        
@bot.message_handler(commands=['help'])
def help(message):
    try:
        text = ""
        if(bot_utils_static.is_group_chat(message)):
            text = help_message_for_group        
        else:
            if(bot_utils_static.is_not_admin(message, comands_admins)):# admin
                text = help_message_in_bot_for_user
            else:
                text = help_message_in_bot_for_admin
        bot.send_message(message.chat.id, text)
    except (Exception)as error:
        print(f'{datetime.today()} ERROR in help()   description: {error}  ')


# Командя для бота (не работает в группе), которая создает пригласительные ссылки для групп где присутствует бот.
@bot.message_handler(commands=["invitation","приглашение"])
@data_handler
def get_invitation_for_group(cursor,message):
    try:       
        if(bot_utils_static.is_not_admin(message, comands_admins) or bot_utils_static.is_group_chat(message)):
            bot.delete_message(message.chat.id, message.message_id)
            return False  
        else:        
            group_list = db_utils_static.select_all_user_chat_group(cursor)
            list=[]
            text=''
            for group in group_list:             
                if(group[1] not in  list):
                    list.append(group[1]) 
                    chat = bot.get_chat(group[1])
                    invite_link = f'{bot_url}?start=+{group[1]}'
                    text = f'''Приглашение в группу <{chat.title}>  link {invite_link}
                     '''
                    bot.send_message(message.chat.id, text)
            return True
    except (Exception)as error:
        print(f'error in get_invitation_for_group(): {datetime.today()} description: {error}  ')



# Командя для бота (не работает в группе), которая создает Excel файл. На первом листе файла все пользователи бота с необходимой информацией. Следующие листы создаются с теми группами где есть бот и списком пользователей группы.    
@bot.message_handler(commands=['list_user_excel'])
@data_handler
def list_user_excel(cursor,message):
    try: 
        users_list = db_utils_static.select_all_users(cursor)
        #create exel        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = f'list_user'        
        users_list = db_utils_static.select_all_users(cursor)
        bot_utils_static.add_header_worksheet_excel_buffer(sheet)
        row_index = 2
        for user in users_list:
             bot_utils_static.add_row_worksheet_excel_buffer(sheet, row_index, user)
             row_index+=1             
       

        # all groups               
        sheet_index = 1
        group_list = db_utils_static.select_all_user_chat_group(cursor)
        list=[]
        for group in group_list:
            row_index = 1
            if(group[1] not in  list):
                chat = bot.get_chat(group[1])
                sheet = wb.create_sheet(f'{chat.title[:30]}',sheet_index)                
                wsheet_index += 1

                users_list = db_utils_static.select_user_by_chat_id(cursor, group[1])
                bot_utils_static.add_header_worksheet_excel_buffer(sheet)  # row_index = 1 
                
                row_index = 2               
                for user in users_list:
                    bot_utils_static.add_row_worksheet_excel_buffer(sheet, row_index, user)
                    row_index+=1

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        # Отправляем файл в чат
        bot.send_document(message.chat.id, ('Users.xlsx', excel_buffer))
    except Exception as error:
        print(f'{datetime.today()} ERROR in list_user_excel()   description: {error}  ')



@bot.message_handler(content_types=["new_chat_members"])
@data_handler
def handler_new_member(cursor, message):
    print("new member")
    try:
        new_chat_members = message.new_chat_members
        for new_member in new_chat_members:
            telegram_user_id = new_member.id
            print(new_member)
       
            user = User(telegram_user_id, User_status.not_confirmed.value, new_member.is_bot)
            current_user = db_utils_static.check_user_exists(cursor, user) 
            user_chat_group = User_chat_group(telegram_user_id, message.chat.id)
            db_utils_static.check_user_chat_group_exists(cursor, user_chat_group)
    
            if current_user.status != User_status.confirmed.value and current_user.is_bot == False and current_user.is_on_service == False:
                bot_utils_static.create_button_auth(telegram_user_id, sso_params, bot,message.chat.id)
                if(bot_utils_static.is_group_chat(message)):
                    text = f'''{new_member.username} {message_text}'''
                    bot.send_message(message.chat.id, text)       
            if (current_user.status == User_status.confirmed.value):    
                bot_utils_static.set_chat_administrator_custom_title(cursor, user, bot)
    except Exception as error:
         print(f'{datetime.today()} ERROR in handler_new_member()   description: {error}  ')

@bot.message_handler(content_types=["left_chat_member"])
@data_handler
def handler_left_member(cursor,message):
    print("left_chat_member")
    try:
        if(message.left_chat_member.id in bot_id):    
            db_utils_static.delete_chat_group(cursor,message.chat.id)
        else:
            db_utils_static.delete_user_from_chat_group(cursor,message.chat.id,message.left_chat_member.id )
    except Exception as error:
         print(f'{datetime.today()} ERROR in handler_left_member()   description: {error}  ')

@bot.message_handler(content_types=["text"] )
@data_handler
def echo(cursor, message): # проверить авторизацию  
    #print(f'Echo пользователь {message.from_user.id } sey {message.text} in {message.chat.id}')     
    try:
        user_writer = db_utils_static.select_user_by_id(cursor, message.from_user.id )
        if(user_writer == False):
                user_writer = User(message.from_user.id, User_status.not_confirmed.value, message.from_user.is_bot)
                db_utils_static.insert_user_db(cursor,user_writer)
        if(bot_utils_static.is_group_chat(message)):  
            user_chat_group = User_chat_group(message.from_user.id, message.chat.id)
            db_utils_static.check_user_chat_group_exists(cursor, user_chat_group)
        if(user_writer.is_on_service==True):
                return True
        if(user_writer.is_bot == True):
                return True    
        #!!! Нужно проверить авторизовался пользователь или нет
        if(user_writer.status != User_status.confirmed.value):
            if(bot_utils_static.is_private_chat(message)):
                bot_utils_static.create_button_auth(user_writer.telegram_user_id, sso_params, bot)
                return True 
            else:
                bot.delete_message(message.chat.id, message.message_id)                    
                text = f'''@{message.from_user.username} {message_text}?start=+{message.chat.id})'''
                bot.send_message(message.chat.id, text)
                bot_utils_static.create_button_auth(user_writer.telegram_user_id, sso_params, bot)
        #!!авторизованный пользователь
        else:               
        
            if(bot_utils_static.is_not_admin(message, comands_admins) == False and bot_utils_static.is_private_chat(message)):                     
                if '/promote_on_service' in message.text:                                        
                    bot_utils_static.promote_user_is_on_service(cursor, bot, message, True)
                elif '/downgrade_on_service' in message.text:                            
                    bot_utils_static.promote_user_is_on_service(cursor, bot, message, False)
                elif '/unban'  in message.text:
                    bot_utils_static.unban_user(cursor, bot, message)
 
            #print(f'!Авторизованный пользователь {message.from_user.id } sey {message.text} in {message.chat.id}' )   
            return True
    except Exception as error:
         print(f'{datetime.today()} ERROR in echo()   description: {error}  ')
    




#для web sever flask
#Запускаем FlaskApp
def starFlaskApp():
    print("Запускаем FlaskApp")
    from waitress import serve
    serve(app, host = host_flask, port = port_flask)
    #app.run()  

#home
@app.route("/")
def home():     
    return f'''<script src="https://bot.jaicp.com/chatwidget/pvzCEFFu:35596d1bbc69ae8b9ed09cfc7bb4c672fb70eabb/justwidget.js?force=true" async> </script>'''
    return "<a>TMK</a>"


# Обработчик для редиректа после аутентификации
@app.route('/callback', methods=['GET'])
@data_handler
def callback(cursor):
    #http_error = "<table> "
    print("callback")
    try:
        # Получаем код аутентификации из параметров запроса
        code = request.args.get('code') 
        state_data = request.args.get('state')
        if((state_data is None ) or (code is None)):
            return "<a>Ошибка получения токена доступа.</a>"
        telegram_user_id=state_data
        telegram_group_id=""
        if("-" in state_data):
            x = state_data.split("-")
            telegram_user_id = x[0]
            telegram_group_id = f'-{x[1]}'
        response = bot_utils_static.get_access_token(code, sso_params)
        # Парсим ответ и извлекаем токен доступа
        if response.status_code == 200:
            token = response.json().get("access_token")
            expires_in = response.json().get("expires_in")        
            refresh_token = response.json().get("refresh_token")             
            userinfo_response = requests.get(userinfo_endpoint, headers={'Authorization': f'Bearer {token}'})
            if userinfo_response.status_code == 200:
                #{'aud': '1ef2af2c-f44d-4aa1-9d72-745e9a5dd179', 
                #'employeeguid': '12231529', 
                #'hrstate': 'Работает', 
                #'insurance': 'F4E333345555B807CAB63B3D1D01169124', 
                #'name': 'ssss ssss sss', 
                #'sub': 'sssss-e7c7-41db-a044-b8d49c141647', 
                #'upn': 'sssss@tmk-group.com'}
                user_name = userinfo_response.json().get("name")  
                #user_hrstate = userinfo_response.json().get("hrstate") 
                employeeguid = userinfo_response.json().get("employeeguid")               

                               
                user = User(telegram_user_id, User_status.confirmed.value, False, user_name, token, expires_in, refresh_token)
                user.employeeguid=employeeguid                
                user_exists = db_utils_static.select_user_by_id(cursor, user.telegram_user_id)                          
                if ((user_exists == False) or ( user_exists is None) ):                   
                    db_utils_static.insert_user_db(cursor, user)
                else:
                    if (user_exists.status == User_status.baned.value or user_exists.status == User_status.baned_for_ever.value):
                        try:                            
                            db_utils_static.update_user_db(cursor, user)
                            bot_utils_static.unban_user_in_all_groups(cursor, bot, user)                             
                        except (Exception) as error:
                            print(f'/callback bot_utils_static.unban_user_in_all_groups { error}')                 
                    db_utils_static.update_user_db(cursor, user)   
                   

                #print(userinfo_data)
                bot.send_message(telegram_user_id, f'''Здравствуйте {user_name}!
                Вы успешно прошли аутентификацию.''')
                bot_utils_static.set_chat_administrator_custom_title(cursor, user, bot)             

                #invitacion 
                if(telegram_group_id != ""):                
                    user_chat_group = User_chat_group(telegram_user_id, telegram_group_id)
                    try:
                        user_telegram = bot.get_chat_member(telegram_group_id,telegram_user_id)                    
                        if((user_telegram is not None) and ( user_telegram.status != 'creator')):     
                            bot.unban_chat_member(telegram_group_id, user.telegram_user_id)
                    except (Exception) as error:
                        print(f'/callback #invitacion { error}')                 
                    db_utils_static.check_user_chat_group_exists(cursor, user_chat_group)
                    link_invite = bot.create_chat_invite_link(telegram_group_id)## depende del grupo que se quiere
                    invite_link = link_invite.invite_link
                    bot.send_message(telegram_user_id, f'''{invite_link}''')
                #/invitacion 
               
            else:
                print("Ошибка получения токена доступа.")
                print(userinfo_response)
                return "<a>Ошибка получения токена доступа.</a>" #+ http_error
        else:
            print("Ошибка получения токена доступа.")
            print(response.text)
            return "<a>Ошибка получения токена доступа.</a>" #+ http_error  
    except Exception as error:
         print(f'{datetime.today()} ERROR in callback()   description: {error}  ')
         return "<a>Ошибка получения токена доступа.</a>" #+ http_error
    return redirect(bot_url)#delet


#автоматизированная система очистки
@data_handler
def cleaning(cursor): 
   #create exel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f'excel_error_report'
    # Заполняем ячейки в примере    
    bot_utils_static.add_header_worksheet_error_report(sheet)     
    row_index = 2

    user_records = db_utils_static.select_all_users_with_group_and_not_duty(cursor)
    for row in user_records:            
        try:
            user = bot_utils_static.get_user_from_record(row)            
            if user.is_bot == True:
                continue
            if user.is_on_service == True:
                continue 
            if(user.status == User_status.baned_for_ever.value or user.status == User_status.baned.value):
                continue        
            if(user.employeeguid is None) or (user.employeeguid==""):
                response_refresh_token = bot_utils_static.get_refresh_token(user.refresh_token, sso_params)
                if response_refresh_token.status_code == 200:
                    user.employee_token = response_refresh_token.json().get("access_token")
                    user.expires_in = response_refresh_token.json().get("expires_in")
                    user.refresh_token = response_refresh_token.json().get("refresh_token") 
                    userinfo_response = requests.get(userinfo_endpoint, headers={'Authorization': f'Bearer {user.employee_token}'})
                    if (userinfo_response.status_code == 200): # it is ok
                        user.user_name = userinfo_response.json().get("name")            
                        user.employeeguid = userinfo_response.json().get("employeeguid")
                    else: # it is error
                    ############### loguin ##################
                        report_text="clean conection error userinfo_response"
                        row_index +=1
                        bot_utils_static.add_row_worksheet_error_report(sheet, row_index, user,report_text)
                        continue                         
                else:    # it is error
                    report_text="clean conection error response_refresh_token"
                    row_index +=1
                    bot_utils_static.add_row_worksheet_error_report(sheet, row_index, user,report_text)
                    ############### loguin ##################              
                    continue 
               
                                           
                
            #     esto es para el nuevo sistema de confirmacion   
            employee = Employee()
            employee.fired = "false"
            try:
                data_requests={ 'token':token_get_employee, 'employeeGUID':user.employeeguid}                
                response_requests = requests.get(url_get_employee,data_requests )                
                if (response_requests.status_code == 200):                    
                    employee = Employee(response_requests.text)                      
                else: # report it is error                    
                    user.status = User_status.not_confirmed.value
                    report_text = f'response_requests status: {response_requests.status_code}'                    
                    bot_utils_static.add_row_worksheet_error_report(sheet, row_index, user,report_text) 
                    row_index +=1
                    continue
            except (Exception) as error:  #pedir loguin             
                user.status = User_status.not_confirmed.value
                print(f'{datetime.today()} ERROR in cleaning();data_xml_as_string;  {user.telegram_user_id}  description: {error}  ')  
                report_text = f'{error}'                
                bot_utils_static.add_row_worksheet_error_report(sheet, row_index, user,report_text)
                row_index +=1
                continue            
            if(employee.fired == "true"): #ban user for ever
                bot_utils_static.ban_user(cursor, bot,user)
                user.status = User_status.baned_for_ever.value
            else:
                user.status = User_status.confirmed.value         
            
        except (Exception) as error:
            print(f'{datetime.today()} ERROR in cleaning()  {user.telegram_user_id}  description: {error}  ')
        finally:    
            #print("final")
            db_utils_static.update_user_db(cursor, user)
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    if(row_index > 2):    
        for admin_Id in comands_admins:
            # Отправляем файл в чат
            bot.send_document(admin_Id, ('Cleaning reort.xlsx', excel_buffer))
    
    
       
        
  
def cleaning_loop():
    print("Запускаем cleaning_loop")   
    while True:        
        try:           
            sleep(2) 
            print(f'cleaning {datetime.today().date()}')
            #cleaning()        
            sleep(cleaning_period) # 1 day
        except (Exception)as error:
            print(f'{datetime.today()} ERROR in cleaning_loop():  description: {error}  ')
            continue
        
        

   
#start
threading.Thread(target = starFlaskApp).start()
threading.Thread(target = starBot).start()
threading.Thread(target = cleaning_loop).start()