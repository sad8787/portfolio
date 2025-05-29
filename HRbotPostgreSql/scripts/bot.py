import os
from test import start_test_block, handle_choice, send_question, \
    test_is_active, handle_text_answer, question_is_open
from contacts import send_contact_request, set_contacts, user_is_leaving_contacts
from decorators import data_handler
from telebot import TeleBot, types
import openpyxl
import re
import traceback
import configparser
import psycopg2


import xlsxwriter
import asyncio
import threading
from time import sleep
import sys
from datetime import date, datetime
from datetime import timedelta


dir_path = os.path.dirname(os.path.realpath(__file__))

config = configparser.ConfigParser()
config.read(f'{dir_path}//..//data//config.ini')


token = os.environ.get('token')
#token = config['bot']['token']

bot = TeleBot(token)


#bd
user_OS = os.environ.get('user')
password_OS = os.environ.get('password')

#user_OS = config['bd']['user']
#password_OS = config['bd']['password']
host_OS = config['bd']['host']
port_OS = config['bd']['port']
database_OS = config['bd']['database']

connection = psycopg2.connect(
        user = user_OS,
        password = password_OS ,
        host = host_OS,
        port = port_OS,
        database = database_OS
       ) 




# Обработчик команды /start
@bot.message_handler(commands=['start'])
@data_handler
def start(cursor, message):    
    admins_ids = [1622663308, 1371591132,1555620161]
    
    if message.from_user.id not in admins_ids:
        check_if_user_exists = f'SELECT user_id FROM users_contacts WHERE user_id = {message.from_user.id}'
        cursor.execute(check_if_user_exists)
        user_exists = cursor.fetchone()

        if user_exists is not None:
            return

    set_contacts(message.from_user.id, message.from_user.username, True)

    keyboard = types.InlineKeyboardMarkup()

    start_button = types.InlineKeyboardButton('Согласен', callback_data='leave_contacts')
    keyboard.add(start_button)

    bot_image_path = f'{dir_path}//..//resources//томка1.png'
    with open(bot_image_path, 'rb') as image:
        bot.send_photo(message.chat.id, image.read())

    start_message = 'Я чат-бот Томка, помогу определить, насколько прокачено у тебя логическое мышление,' \
                    ' ведь на большинстве ИТ вакансий оно пригодится. Если ответишь правильно на все вопросы,' \
                    ' получишь приз. Также я познакомлю тебя с актуальными позициями в ТМК++ и буду держать кулачки,' \
                    ' чтобы ты стал частью нашей команды.\nДля продолжения общения, ' \
                    'необходимо твое согласие на обработку персональных данных, если не возражаешь,' \
                    ' то ознакомься с текстом согласия' \
                    ' и после этого нажми на кнопку «Согласен». ' \
                    'Текст согласия ты сможешь найти во вложениях ниже:'

    bot.send_message(message.chat.id, start_message, reply_markup=keyboard)

    agreement_path = f'{dir_path}//..//resources//согласие.docx'
    with open(agreement_path, 'rb') as doc:
        bot.send_document(
            message.chat.id,
            doc.read(),
            visible_file_name='Согласие на обработку персональных данных.docx'
        )


# Обработчик команды /data
@bot.message_handler(commands=['supersecretloadoutcommandthatisknownonlybyme'])
@data_handler
def get_test_results(cursor, message):

    print(f'supersecretloadoutcommandthatisknownonlybyme ')
    print("")
    
    try:
        dir_excel=f'{dir_path}//..//tables//Результаты тестирования2.xlsx'
        workbook = xlsxwriter.Workbook(dir_excel)
        worksheet = workbook.add_worksheet(f'Участники')
        worksheet.write(0,0,"Telegram")
        worksheet.write(0,1,"Правильных ответов")
        worksheet.write(0,2,"ФИО")  
        worksheet.write(0,3,"Дата рождения")  
        worksheet.write(0,4,"ВУЗ")
        worksheet.write(0,5,"Квалификация")
        worksheet.write(0,6,"Курс")  
        worksheet.write(0,7,"Факультет")  
        worksheet.write(0,8,"Специальность")
        worksheet.write(0,9,"Номер телефона")
        worksheet.write(0,10,"Сфера IT")
        


        exel_row_count=0
        exel_column =0
        cursor = connection.cursor()
        get_all_user=f'''SELECT user_id, is_test_active, block_id, questions_order, question_index, contact_index, agreement, entry_date FROM users;'''
        cursor.execute(get_all_user)
        all_user = cursor.fetchall()
        index=1
        for row_user in all_user:
            get_all_contacts = f'SELECT user_id, contact_name, contact_value FROM users_contacts WHERE user_id={row_user[0]} '
            cursor.execute(get_all_contacts)
            all_contacts = cursor.fetchall()
            for row_contacts in all_contacts:
                if(row_contacts[1]=="Telegram"):
                     worksheet.write(index,0,f'{row_contacts[2]}')
                if(row_contacts[1]=="ФИО"):
                     worksheet.write(index,2,f'{row_contacts[2]}')
                if(row_contacts[1]=="Дата рождения"):
                     worksheet.write(index,3,f'{row_contacts[2]}')
                if(row_contacts[1]=="ВУЗ"):
                     worksheet.write(index,4,f'{row_contacts[2]}')
                if(row_contacts[1]=="Квалификация"):
                     worksheet.write(index,5,f'{row_contacts[2]}')                
                if(row_contacts[1]=="Курс"):
                     worksheet.write(index,6,f'{row_contacts[2]}')
                if(row_contacts[1]=="Факультет"):
                     worksheet.write(index,7,f'{row_contacts[2]}')
                if(row_contacts[1]=="Специальность"):
                     worksheet.write(index,8,f'{row_contacts[2]}')
                if(row_contacts[1]=="Номер телефона"):
                     worksheet.write(index,9,f'{row_contacts[2]}')
                if(row_contacts[1]=="Сфера IT"):
                     worksheet.write(index,10,f'{row_contacts[2]}')

            get_count_users_answers= f'''SELECT COUNT (is_correct) FROM users_answers    WHERE user_id={row_user[0]} and is_correct = {1} ;  '''
            cursor.execute(get_count_users_answers)
            count_users_answers=cursor.fetchone()
            worksheet.write(index,1,f'{count_users_answers[0]}')
            index+=1
        
        workbook.close()         
        wb =  open(dir_excel, 'rb')
        bot.send_document(message.chat.id,wb)

    except (Exception, psycopg2.Error) as error:
        print("Error while connecting to PostgreSQL: ", error)
    finally:
        if connection:
            cursor.close()
            connection.close()
            print("PostgreSQL connection is closed")


    return True

    #get_all_contacts = f'SELECT user_id, contact_name, contact_value FROM users_contacts'
    #all_contacts = cursor.execute(get_all_contacts).fetchall()    
    #all_contacts = pd.DataFrame(all_contacts, columns=['user_id', 'contact_name', 'contact_value'])
    #all_contacts_df = pd.DataFrame()

    # Нужно доработать фукнцию!

    #get_all_contacts = f'SELECT user_id, contact_name, contact_value FROM users_contacts'
    #all_contacts = cursor.execute(get_all_contacts).fetchall()
    #all_contacts = pd.DataFrame(all_contacts, columns=['user_id', 'contact_name', 'contact_value'])
    #all_contacts_df = pd.DataFrame()


    #get_users_answers = f'SELECT * from users_answers'
    #users_answers = cursor.execute(get_users_answers).fetchall()
    #users_answers = pd.DataFrame(users_answers, columns=['user_id', 'question_id', 'is_correct'])

    #for user_id, group in all_contacts.groupby('user_id'):
    #    index = len(all_contacts_df)

    #    for row_index, row in group.iterrows():
    #        all_contacts_df.loc[index, row.contact_name] = row.contact_value

    #    correct_answers_count = users_answers[
    #        (users_answers['user_id'] == user_id)
    #        & users_answers['is_correct'] == 1
    #    ].shape[0]

    #    all_contacts_df.loc[index, 'Правильных ответов'] = correct_answers_count

    #table_path = f'{dir_path}//..//tables//Результаты тестирования.xlsx'

    #with pd.ExcelWriter(table_path) as writer:
    #    if len(all_contacts_df) > 0:
    #        all_contacts_df.to_excel(writer, sheet_name='Участники', index=False)

    #wb = openpyxl.load_workbook(filename=table_path)
    #for sheet_name in wb.sheetnames:
    #    worksheet = wb[sheet_name]
    #    for col in worksheet.columns:
    #        max_length = 0
    #        column = col[0].column_letter
    #        for cell in col:
    #            try:
    #                if len(str(cell.value)) > max_length:
    #                    max_length = len(str(cell.value))
    #            except:
    #                pass
    #        adjusted_width = (max_length + 2) * 1.2
    #        worksheet.column_dimensions[column].width = adjusted_width

    #wb.save(table_path)

    #with open(table_path, 'rb') as table:
    #    bot.send_document(message.chat.id, table.read(), visible_file_name=table_path[table_path.rfind('/'):])
    #    table.close()


# Обработчик любого введённого текста
@bot.message_handler(content_types=['text'])
def handle_text(message):
    user_id = str(message.from_user.id)
    chat_id = message.chat.id
    message_text = message.text    
    if user_is_leaving_contacts(user_id):
        message_text = re.sub(r'[^\w\s\.\+\-]', '', message_text)
        if set_contacts(user_id, message_text):
            send_contact_request(chat_id, user_id)
        else:
            bot.send_message(chat_id, 'Неправильный формат, попробуй ещё раз')
    elif test_is_active(user_id) and question_is_open(user_id):
        handle_text_answer(chat_id, user_id, message_text)
        send_question(chat_id, user_id)
    else:
        bot.send_message(chat_id, message_text)


# Обработчик всех кнопок с 'test' в содержимом callbackQuery.data
@bot.callback_query_handler(lambda query: 'test' in query.data)
def handle_test(callback):
    user_id = str(callback.from_user.id)
    chat_id = callback.message.chat.id
    if callback.data == 'test_start':
        bot.delete_message(chat_id, callback.message.id)        
        if test_is_active(user_id):
            bot.send_message(chat_id, 'Тест уже идёт')
        else:
            start_test_block(user_id)
            send_question(chat_id, user_id)
    elif callback.data == 'test_continue':
        bot.delete_message(chat_id, callback.message.id)
        start_test_block(user_id)
        send_question(chat_id, user_id)
    else:
        handle_choice(chat_id, user_id, callback)
        send_question(chat_id, user_id)


# Обработчик всех кнопок с 'contacts' в содержимом callbackQuery.data
@bot.callback_query_handler(lambda query: 'contacts' in query.data)
@data_handler
def handle_contacts(cursor, callback):
    bot.answer_callback_query(callback.id)
    user_id = str(callback.from_user.id)
    chat_id = callback.message.chat.id
    if callback.data == 'leave_contacts':
        send_contact_request(chat_id, user_id, True)







def database_cleaner():
    print ("database_cleaner")    
    connection = psycopg2.connect(
        user = user_OS,
        password = password_OS ,
        host = host_OS,
        port = port_OS,
        database = database_OS
        ) 

@data_handler
def database_cleaner(cursor):
    print ("database_cleaner")
    #connection = psycopg2.connect(
    #    user = config['bd']['user'],
    #    password = config['bd']['password'],
    #    host=config['bd']['host'],
    #    port=config['bd']['port'],
    #    database=config['bd']['database']
    #    )

    try:
        #cursor = connection.cursor()
        today = datetime.today().date()      
        day_delta=timedelta(days=60)
        old_day=today-day_delta
        value_date = old_day.strftime('%Y-%m-%d')
        print(value_date)   
            #delete_user ="DELETE FROM users WHERE entry_date <=' "+value+"';"
            #INSERT INTO public.users(	user_id, entry_date) VALUES (15778887, '2023-01-01');
            #INSERT INTO public.users_contacts(	user_id, contact_name, contact_value) VALUES (1, 'mail', 's@de.com');
            #INSERT INTO public.users_answers(user_id, question_id, is_correct)	VALUES (1, 3, 0);
            #INSERT INTO public.users_open_answers(	user_id, question_id, answer)	VALUES (5, 1, 1);
        get_user_id = f'''SELECT user_id FROM users WHERE entry_date <= '{value_date}';'''
        cursor.execute(get_user_id)
        users_id = cursor.fetchall()
        print("----")
        print(users_id)
        print("----")
        for row in users_id:
            delete_users_answers = "DELETE FROM users_answers WHERE user_id = "+str(row[0])+";"
            delete_users_contacts ="DELETE FROM users_contacts WHERE user_id = "+str(+row[0])+";"
            delete_users_open_answers ="DELETE FROM users_open_answers	WHERE user_id = "+str(+row[0])+";"
            delete_user = "DELETE FROM users WHERE user_id = '"+str(+row[0])+"';"
            delete_data = delete_users_answers+delete_users_contacts+delete_users_open_answers+delete_user
            cursor.execute(delete_data)
    
            #cursor.execute(delete_user)
    
            #connection.commit()
    except (Exception, psycopg2.Error) as error:
        print("Error while connecting to PostgreSQL: ", error)
    #finally:
    #    if connection:
    #        cursor.close()
    #        connection.close()
    #        print("PostgreSQL connection is closed")

def database_cleaner_loop():
    while True:
        print("database_cleaner_loop is running")        
        sleep(86400)
        database_cleaner()




    
#threading.Thread(target=starBot).start()
threading.Thread(target=database_cleaner_loop).start()
# Запуск бота
if __name__ == '__main__':    
    while True:
        try:
           print("Запускаем бота")
           bot.polling(none_stop=True, interval=0)           
        except Exception:
            print(traceback.format_exc())
            continue
